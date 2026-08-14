import os
import zipfile
import sqlite3
import hashlib
import glob
from datetime import datetime
import xml.etree.ElementTree as ET
import openpyxl

# Connect to database
db_path = os.path.join("prisma", "dev.db")
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

def sha256_hash(text):
    return hashlib.sha256(text.encode()).hexdigest()

def get_now_iso():
    return datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.000Z')

print("Starting data seeding process...")

# 1. Create Default Users
users = [
    ("Super Admin", "admin@askarisolar.com", sha256_hash("admin123"), "Admin", "Management"),
    ("HR Manager", "hr@askarisolar.com", sha256_hash("hr123"), "HR", "HR"),
    ("Accounts Accountant", "accounts@askarisolar.com", sha256_hash("accounts123"), "Accountant", "Accounts"),
    ("Sales Specialist", "sales@askarisolar.com", sha256_hash("sales123"), "Sales & Marketing Department", "Sales"),
    ("Field Staff Worker", "field@askarisolar.com", sha256_hash("field123"), "Field Staff", "Field")
]

cursor.execute("DELETE FROM User")
for name, email, password, role, dept in users:
    cursor.execute(
        "INSERT INTO User (name, email, password, role, department, createdAt, updatedAt) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (name, email, password, role, dept, get_now_iso(), get_now_iso())
    )
conn.commit()
print("Seeded default users.")

# Get primary users IDs for file uploads
cursor.execute("SELECT id, role FROM User")
user_map = {role: uid for uid, role in cursor.fetchall()}
admin_uid = user_map.get("Admin", 1)
sales_uid = user_map.get("Sales & Marketing Department", 1)
accounts_uid = user_map.get("Accountant", 1)
hr_uid = user_map.get("HR", 1)

# 2. Extract ZIP Files
upload_root = os.path.join("public", "uploads")
os.makedirs(upload_root, exist_ok=True)

data_folder = "_data"
zip_files = glob.glob(os.path.join(data_folder, "*.zip"))

for zip_path in zip_files:
    zip_name = os.path.basename(zip_path)
    print(f"Extracting {zip_name}...")
    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(upload_root)
    except Exception as e:
        print(f"Error unzipping {zip_name}: {e}")

print("ZIP files extracted to public/uploads.")

# 3. Parse Lectures Links (docx)
lectures_docx = os.path.join(upload_root, "Lectures Links", "Untitled document.docx")
lectures = []
if os.path.exists(lectures_docx):
    try:
        with zipfile.ZipFile(lectures_docx, 'r') as docx_zip:
            doc_xml = docx_zip.read('word/document.xml')
            root = ET.fromstring(doc_xml)
            ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
            
            p_texts = []
            for p in root.findall('.//w:p', ns):
                p_text = "".join(t.text for t in p.findall('.//w:t', ns) if t.text)
                if p_text.strip():
                    p_texts.append(p_text.strip())
            
            # Combine Title + YouTube Link
            current_title = "General Solar Training"
            for i, text in enumerate(p_texts):
                if "Link:" in text or "http" in text:
                    url = text.split("Link:")[-1].strip() if "Link:" in text else text.strip()
                    if url.startswith("F "):
                        url = url[2:]
                    lectures.append((current_title, url))
                else:
                    if len(text) < 100:
                        current_title = text
    except Exception as e:
        print(f"Error parsing training lectures docx: {e}")
else:
    print("Lectures docx not found at path:", lectures_docx)

# 4. Process FileItem table structure from extracted directories
cursor.execute("DELETE FROM FileItem")
conn.commit()

# Helper function to recursively populate files into Database
def seed_folder_files(dir_path, dept, doc_type, parent_id=None, virtual_base_path=""):
    if not os.path.exists(dir_path):
        return
    
    entries = os.listdir(dir_path)
    for entry in entries:
        full_path = os.path.join(dir_path, entry)
        rel_url_path = os.path.relpath(full_path, "public").replace("\\", "/")
        
        is_folder = os.path.isdir(full_path)
        ext = os.path.splitext(entry)[1].lower().replace(".", "") if not is_folder else None
        size = os.path.getsize(full_path) if not is_folder else None
        
        color = None
        if is_folder:
            # Assign cute colors
            colors = ["amber", "blue", "emerald", "violet", "rose", "indigo"]
            color = colors[len(entry) % len(colors)]
        
        cursor.execute(
            """INSERT INTO FileItem 
               (name, isFolder, folderColor, path, parentId, department, docType, fileExtension, fileSize, fileUrl, uploadedById, isFavorite, version, createdAt, updatedAt)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                entry, is_folder, color, virtual_base_path, parent_id,
                dept, doc_type, ext, size, f"/{rel_url_path}", sales_uid,
                False, 1, get_now_iso(), get_now_iso()
            )
        )
        entry_id = cursor.lastrowid
        
        if is_folder:
            seed_folder_files(full_path, dept, doc_type, entry_id, f"{virtual_base_path}/{entry}")

# Let's seed folders from extracted zips
print("Seeding File Manager database...")

# Sales Advertisements
seed_folder_files(os.path.join(upload_root, "Advertisement Data"), "Sales", "Advertisements", None, "/Advertisements")

# Accounts Training
seed_folder_files(os.path.join(upload_root, "Training (Accounts Department)"), "Accounts", "Training", None, "/Training/Accounts")

# HR Legal Documents
seed_folder_files(os.path.join(upload_root, "Documents"), "HR", "Documents", None, "/HR/Documents")

# Lectures folder and youtube links
cursor.execute(
    """INSERT INTO FileItem 
       (name, isFolder, folderColor, path, parentId, department, docType, createdAt, updatedAt)
       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
    ("Lectures Links & Videos", True, "violet", "/Training/Sales", None, "Sales", "Training", get_now_iso(), get_now_iso())
)
sales_tr_folder_id = cursor.lastrowid

for title, url in lectures:
    cursor.execute(
        """INSERT INTO FileItem 
           (name, isFolder, path, parentId, department, docType, fileExtension, fileUrl, uploadedById, createdAt, updatedAt)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            title, False, "/Training/Sales/Lectures Links & Videos", sales_tr_folder_id,
            "Sales", "Training", "link", url, sales_uid, get_now_iso(), get_now_iso()
        )
    )

# Quotations Folder for Islamabad & Chakwal
cursor.execute(
    """INSERT INTO FileItem 
       (name, isFolder, folderColor, path, parentId, department, docType, createdAt, updatedAt)
       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
    ("Quotations", True, "emerald", "/Quotations", None, "Accounts", "Quotations", get_now_iso(), get_now_iso())
)
quot_folder_id = cursor.lastrowid

# Islamabad Quotations
cursor.execute(
    """INSERT INTO FileItem 
       (name, isFolder, folderColor, path, parentId, department, docType, createdAt, updatedAt)
       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
    ("Islamabad Office", True, "blue", "/Quotations/Islamabad Office", quot_folder_id, "Accounts", "Quotations", get_now_iso(), get_now_iso())
)
isl_folder_id = cursor.lastrowid
seed_folder_files(os.path.join(upload_root, "Islamabad"), "Accounts", "Quotations", isl_folder_id, "/Quotations/Islamabad Office")

# Chakwal Quotations
cursor.execute(
    """INSERT INTO FileItem 
       (name, isFolder, folderColor, path, parentId, department, docType, createdAt, updatedAt)
       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
    ("Chakwal Office", True, "amber", "/Quotations/Chakwal Office", quot_folder_id, "Accounts", "Quotations", get_now_iso(), get_now_iso())
)
chk_folder_id = cursor.lastrowid
seed_folder_files(os.path.join(upload_root, "Chakwal"), "Accounts", "Quotations", chk_folder_id, "/Quotations/Chakwal Office")

conn.commit()
print("Seeded FileManager and folder structures.")

# 5. Parse Price Lists Excel and Seed Product Database
print("Parsing Excel Sheet & Seeding Products...")
cursor.execute("DELETE FROM Product")
conn.commit()

excel_path = os.path.join("_data", "Price Lists.xlsx")
if os.path.exists(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # We will parse: Yinergy, Inverex, Crown, Knox, DIPOWER Batteries, Flate & Tubular Batteries, Cables, Protection Devices, Suntrix
        # Yinergy
        if 'Yinergy' in wb.sheetnames:
            ws = wb['Yinergy']
            # We see Inverter hybrid / offgrid, let's extract them
            for row in ws.iter_rows(min_row=5, max_row=40, values_only=True):
                if row and len(row) >= 9 and row[0] == "Inverter":
                    name = f"Yinergy {row[3]} Inverter ({row[1]})"
                    spec = f"Power: {row[3]}, PV: {row[4]}, Wifi: {row[5]}"
                    try:
                        rate = float(row[6]) if isinstance(row[6], (int, float)) else 55000.0
                    except:
                        rate = 55000.0
                    cursor.execute(
                        "INSERT INTO Product (category, name, brand, spec, rate, purchasePrice, stock, warranty, createdAt, updatedAt) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        ("Inverters", name, "Yinergy", spec, rate, rate * 0.85, 25, "5 Years", get_now_iso(), get_now_iso())
                    )
        
        # Inverex
        if 'Inverex' in wb.sheetnames:
            ws = wb['Inverex']
            for row in ws.iter_rows(min_row=6, max_row=30, values_only=True):
                if row and len(row) >= 4 and row[2] is not None:
                    name = f"Inverex {row[2]} Inverter"
                    spec = f"Phase: {row[4]}, IP: {row[5]}" if len(row) >= 6 else ""
                    rate_str = str(row[6]) if len(row) >= 7 else "300"
                    # price is like "235k"
                    rate = 250000.0
                    if 'k' in rate_str.lower():
                        try:
                            rate = float(rate_str.lower().replace('k', '').strip()) * 1000
                        except:
                            pass
                    cursor.execute(
                        "INSERT INTO Product (category, name, brand, spec, rate, purchasePrice, stock, warranty, createdAt, updatedAt) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        ("Inverters", name, "Inverex", spec, rate * 1.15, rate, 18, "5 Years", get_now_iso(), get_now_iso())
                    )
                    
        # Crown
        if 'Crown' in wb.sheetnames:
            ws = wb['Crown']
            for row in ws.iter_rows(min_row=3, max_row=25, values_only=True):
                if row and len(row) >= 5 and row[2] is not None:
                    name = f"Crown {row[2]} {row[0]} Inverter"
                    spec = f"IP rating: {row[1]}"
                    try:
                        purchase = float(row[3])
                        sale = float(row[4])
                    except:
                        purchase = 100000.0
                        sale = 120000.0
                    cursor.execute(
                        "INSERT INTO Product (category, name, brand, spec, rate, purchasePrice, stock, warranty, createdAt, updatedAt) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        ("Inverters", name, "Crown Micro", spec, sale, purchase, 15, "3 Years", get_now_iso(), get_now_iso())
                    )

        # DIPOWER Batteries
        # Let's seed DIPOWER Batteries with default structured data
        dipower_batts = [
            ("DIPOWER Lithium Battery 100Ah 48V", "Lithium Battery", 420000.0, 390000.0, 10, "5 Years Warranty"),
            ("DIPOWER Lithium Battery 200Ah 48V", "Lithium Battery", 780000.0, 720000.0, 5, "5 Years Warranty"),
            ("DIPOWER Tubular Battery 180Ah", "Tubular Battery", 55000.0, 48000.0, 40, "1 Year Warranty"),
            ("DIPOWER Tubular Battery 220Ah", "Tubular Battery", 65000.0, 58000.0, 30, "1 Year Warranty")
        ]
        for name, spec, sale, purchase, stock, war in dipower_batts:
            cursor.execute(
                "INSERT INTO Product (category, name, brand, spec, rate, purchasePrice, stock, warranty, createdAt, updatedAt) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                ("Batteries", name, "DIPOWER", spec, sale, purchase, stock, war, get_now_iso(), get_now_iso())
            )

        # Cables
        if 'Cables' in wb.sheetnames:
            ws = wb['Cables']
            for row in ws.iter_rows(min_row=3, max_row=30, values_only=True):
                if row and len(row) >= 8 and row[2] is not None:
                    item_name = f"AC Cable {row[2]} {row[4] or ''}"
                    cursor.execute(
                        "INSERT INTO Product (category, name, brand, spec, rate, purchasePrice, stock, warranty, createdAt, updatedAt) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        ("Cables", item_name, "GM/Newage", row[3] or "Standard", float(row[7]) if isinstance(row[7], (int,float)) else 6000.0, float(row[6]) if isinstance(row[6], (int,float)) else 5000.0, 100, "10 Years", get_now_iso(), get_now_iso())
                    )

        # Protection Devices
        if 'Protection Devices' in wb.sheetnames:
            ws = wb['Protection Devices']
            for row in ws.iter_rows(min_row=3, max_row=30, values_only=True):
                if row and len(row) >= 8 and row[2] is not None:
                    acc_name = f"AC Protection {row[2]} {row[3] or ''} {row[4] or ''}"
                    cursor.execute(
                        "INSERT INTO Product (category, name, brand, spec, rate, purchasePrice, stock, warranty, createdAt, updatedAt) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        ("Accessories", acc_name, row[5] or "CNC", "Protective Device", float(row[7]) if isinstance(row[7], (int,float)) else 2500.0, float(row[6]) if isinstance(row[6], (int,float)) else 2000.0, 150, "1 Year", get_now_iso(), get_now_iso())
                    )

        # Let's seed Solar Panels (defaults, since not explicit in sheets, but critical for solar energy portal!)
        panels = [
            ("Longi Hi-MO 6 Explorer 575W Mono-Facial", "Longi", "Mono-Perc High Efficiency", 22000.0, 19500.0, 500, "25 Years Performance"),
            ("Jinko Tiger Neo N-type 580W Bifacial", "Jinko Solar", "N-type Dual Glass Bifacial", 24000.0, 21000.0, 450, "30 Years Performance"),
            ("Canadian Solar BiHiKu7 650W Bifacial", "Canadian Solar", "High Power Bifacial Mono", 28000.0, 24500.0, 300, "25 Years Performance")
        ]
        for name, brand, spec, sale, purchase, stock, war in panels:
            cursor.execute(
                "INSERT INTO Product (category, name, brand, spec, rate, purchasePrice, stock, warranty, createdAt, updatedAt) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                ("Solar Panels", name, brand, spec, sale, purchase, stock, war, get_now_iso(), get_now_iso())
            )

        print("Successfully loaded products from Excel sheets.")
    except Exception as e:
        print(f"Error parsing Price Lists Excel: {e}")
else:
    print("Price Lists.xlsx not found.")

# 6. Seed CRM Leads, Customers & Support Tickets
print("Seeding CRM Data...")
cursor.execute("DELETE FROM Lead")
cursor.execute("DELETE FROM Customer")
cursor.execute("DELETE FROM Project")
cursor.execute("DELETE FROM SupportTicket")
conn.commit()

leads = [
    ("Muhammad Usman", "0300-1234567", "35202-1234567-1", "House 12, Phase 4A", "Islamabad", 45000.0, 350, 10.0, "Campaign", "Summer Solar Promo", 3, "New", "Customer interested in 10kW On-Grid solar system.", get_now_iso(), get_now_iso()),
    ("Adeel Sahib", "0321-7654321", "34101-7654321-3", "Main Bazar, Dhudial", "Chakwal", 25000.0, 220, 5.0, "Social", "Facebook Lead Gen", 3, "Contacted", "Called once, scheduled detail callback for weekend.", get_now_iso(), get_now_iso()),
    ("Kamran Khan", "0333-1112223", "37405-1112223-5", "Sector G-11/3", "Islamabad", 85000.0, 680, 15.0, "Direct", None, 3, "Survey Scheduled", "Site survey scheduled for Friday afternoon.", get_now_iso(), get_now_iso()),
    ("Zahid Mehmood", "0345-4445556", None, "Civil Lines", "Chakwal", 15000.0, 150, 3.0, "Reference", "Referred by Muhammad Usman", 3, "Quotation Sent", "Sent quotation for 3.2kW Hybrid system (Yinergy).", get_now_iso(), get_now_iso()),
    ("Sajjad Ali", "0312-9998887", None, "Bahria Town Phase 7", "Islamabad", 62000.0, 500, 12.0, "Campaign", "Google Ads Solar", 3, "Negotiation", "Customer negotiating on final price of 12kW system.", get_now_iso(), get_now_iso()),
    ("Amjad Pervez", "0300-8887776", None, "Chakwal Road", "Chakwal", 35000.0, 300, 8.0, "Direct", None, 3, "Won", "Agreement signed. 8kW Solar system installation starting.", get_now_iso(), get_now_iso())
]

for name, phone, cnic, addr, city, bill, units, ld, src, camp, sp_id, status, notes, cr, up in leads:
    cursor.execute(
        """INSERT INTO Lead 
           (name, phone, cnic, address, city, electricityBill, monthlyUnits, load, source, campaign, salesPersonId, status, notes, createdAt, updatedAt)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (name, phone, cnic, addr, city, bill, units, ld, src, camp, sp_id, status, notes, cr, up)
    )

customers = [
    ("Muhammad Usman", "0300-1234567", "35202-1234567-1", "House 12, Phase 4A", "Islamabad", "usman@gmail.com", get_now_iso(), get_now_iso()),
    ("Amjad Pervez", "0300-8887776", "34301-8887776-5", "Chakwal Road", "Chakwal", "amjad@yahoo.com", get_now_iso(), get_now_iso())
]

for name, phone, cnic, addr, city, email, cr, up in customers:
    cursor.execute(
        "INSERT INTO Customer (name, phone, cnic, address, city, email, createdAt, updatedAt) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (name, phone, cnic, addr, city, email, cr, up)
    )

# Seed Projects matching customer profiles
cursor.execute("SELECT id, name FROM Customer")
cust_map = {name: cid for cid, name in cursor.fetchall()}

projects = [
    (cust_map["Muhammad Usman"], "Usman 10kW On-Grid Project", "Installation", "Site survey completed. Roof is shadow-free.", "Approved by Management", "Material ordered and dispatched.", "2026-08-05", None, "Applied", 5, "Material transit, scheduled setup.", get_now_iso(), get_now_iso()),
    (cust_map["Amjad Pervez"], "Amjad 8kW Hybrid Project", "Completed", "Survey done. Flat concrete roof.", "Approved", "Delivered", "2026-07-20", "2026-07-23", "Approved & Connected", 5, "Completed successfully. Net metering configured.", get_now_iso(), get_now_iso())
]

for cid, name, stage, survey, approval, material, inst_dt, insp_dt, net_met, war, notes, cr, up in projects:
    cursor.execute(
        """INSERT INTO Project 
           (customerId, name, stage, surveyDetails, approvalStatus, materialOrdered, installationDate, inspectionDate, netMeteringStatus, warrantyYears, notes, createdAt, updatedAt)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (cid, name, stage, survey, approval, material, inst_dt, insp_dt, net_met, war, notes, cr, up)
    )

# Support Tickets
tickets = [
    (cust_map["Amjad Pervez"], "Inverter App Syncing Issue", "Customer complaining that the Yinergy Wifi app is not displaying generation data.", "Complaint", "Open", get_now_iso(), get_now_iso()),
    (cust_map["Muhammad Usman"], "Net Metering Status Inquiry", "Usman wants an update on NEPRA net metering billing registration.", "Service Visit", "In Progress", get_now_iso(), get_now_iso())
]
for cid, sub, desc, typ, status, cr, up in tickets:
    cursor.execute(
        "INSERT INTO SupportTicket (customerId, subject, description, type, status, createdAt, updatedAt) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (cid, sub, desc, typ, status, cr, up)
    )

conn.commit()
print("Seeded CRM Leads, Customers, Projects & Tickets.")

# 7. Seed Work Schedule (Based on Excel layout)
print("Seeding Work Schedule...")
cursor.execute("DELETE FROM WorkSchedule")
conn.commit()

schedule_items = [
    ("2026-07-30", "Thursday", "Islamabad", "Dr. Shahid", "0333-5556667", "Sector F-8/2", "Inverter Replacement", "Replace faulty Inverex 10kW inverter under warranty.", 5000.0, "Hamza (Tech) + Ali (Helper)", "2026-07-30", "2026-07-30", "In Progress", "Warranty claim ticket #102.", get_now_iso(), get_now_iso()),
    ("2026-07-31", "Friday", "Chakwal", "Al-Makkah Sweets", "0321-4442221", "Sargodha Road", "Net Metering Inspection", "Liaison with WAPDA inspector for net meter testing.", 8000.0, "Engr. Nabeel", "2026-07-31", "2026-07-31", "Pending", "Verify green meter compatibility first.", get_now_iso(), get_now_iso()),
    ("2026-07-29", "Wednesday", "Islamabad", "Gen. Tariq", "0300-9993332", "DHA Phase 2", "Structure Strengthening", "Add extra cross-bracing to solar panel stands.", 15000.0, "Irshad & Welding Team", "2026-07-29", "2026-07-29", "Completed", "Stands reinforced against high wind speeds.", get_now_iso(), get_now_iso())
]

for dt, day, branch, name, num, loc, nature, desc, chg, staff, start, end, status, rem, cr, up in schedule_items:
    cursor.execute(
        """INSERT INTO WorkSchedule 
           (date, day, branch, clientName, clientNumber, clientLocation, natureOfWork, workDescription, charges, assignedStaff, startDate, endDate, status, remarks, createdAt, updatedAt)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (dt, day, branch, name, num, loc, nature, desc, chg, staff, start, end, status, rem, cr, up)
    )

# 8. Seed Announcements & Attendance Logs
print("Seeding Announcements & Attendance...")
cursor.execute("DELETE FROM Announcement")
cursor.execute("DELETE FROM Attendance")
conn.commit()

announcements = [
    ("New Training Material Uploaded", "Sales and accounts training materials have been updated in the document library. Please review the new procedures.", "All", None, True, admin_uid, get_now_iso(), get_now_iso()),
    ("Target Achievements Q3", "We are close to meeting our sales target for Q3. Management is announcing bonuses for top sales achievers.", "Sales", None, False, admin_uid, get_now_iso(), get_now_iso()),
    ("UBL Bank Account SOPs Change", "All accounts staff must note that the UBL accounts transactions should be verified daily by 4:00 PM. Follow the updated Accounts SOPs file.", "Accounts", "/uploads/Training (Accounts Department)/SOPs.docx", False, admin_uid, get_now_iso(), get_now_iso())
]

for title, content, dept, att, pin, created_by, cr, up in announcements:
    cursor.execute(
        "INSERT INTO Announcement (title, content, department, attachmentUrl, isPinned, createdById, createdAt, updatedAt) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (title, content, dept, att, pin, created_by, cr, up)
    )

# Attendance Logs (Last few days for employee and admin)
attendance_data = [
    (user_map.get("Employee", 7), "2026-07-30", "09:02:15", "18:05:00", "Present", "On time"),
    (user_map.get("Employee", 7), "2026-07-29", "09:15:30", "18:00:00", "Late", "Traffic delay"),
    (user_map.get("Employee", 7), "2026-07-28", "08:58:00", "18:10:00", "Present", "On time"),
    (user_map.get("Sales Agent", 3), "2026-07-30", "08:55:00", "18:00:00", "Present", "On time"),
    (user_map.get("Accounts Officer", 4), "2026-07-30", "09:05:00", "18:02:00", "Present", "On time")
]

for uid, dt, chk_in, chk_out, status, notes in attendance_data:
    cursor.execute(
        "INSERT INTO Attendance (userId, date, checkIn, checkOut, status, notes, createdAt, updatedAt) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (uid, dt, chk_in, chk_out, status, notes, get_now_iso(), get_now_iso())
    )

conn.commit()
conn.close()

print("\nAll database seeding completed successfully! [Done]")
